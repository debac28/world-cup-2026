#!/usr/bin/env python3
"""Extract the static tournament skeleton from fifa_2026.xlsx into public/data/seed.json.

This is a ONE-TIME (or rarely-run) build step. The spreadsheet is a blank template:
it has no live scores and no goal scorers. We only pull the things that never change
during the tournament:

  - the 48 teams (+ ISO codes for flag images)
  - the 12 groups (A-L)
  - the 104 group-stage fixtures, each with a canonical UTC kickoff
  - the full knockout bracket (Round of 32 -> Final, plus 3rd-place play-off),
    expressed as slot rules ("1E", "2A", "3rd Group A/B/C/D/F") and feeder matches.

Live scores / standings / top scorers come from API-Football at runtime
(see scripts/update_results.mjs) and are layered on top of this seed in the app.

The spreadsheet's kickoff times are displayed in GMT+5:30 (the author's configured
timezone). We convert every kickoff to UTC here so the browser can render it in each
user's local timezone unambiguously.
"""
import json
import os
import re
from datetime import timedelta

import openpyxl
from openpyxl.utils.datetime import from_excel

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
XLSX = os.path.join(ROOT, "fifa_2026.xlsx")
OUT = os.path.join(ROOT, "public", "data", "seed.json")

# The spreadsheet's Settings sheet has GMT-Time = "GMT + 5:00" and Minutes = "+30 min".
# Displayed kickoff = UTC + 5:30, so UTC = displayed - 5:30.
SHEET_TZ_OFFSET = timedelta(hours=5, minutes=30)

# Team name -> ISO code understood by flagcdn.com (https://flagcdn.com/w80/<code>.png).
# Non-country / sub-national entries use flagcdn's supported codes.
FLAG_CODES = {
    "Algeria": "dz", "Argentina": "ar", "Australia": "au", "Austria": "at",
    "Belgium": "be", "Bosnia and Herzegovina": "ba", "Brazil": "br", "Canada": "ca",
    "Cape Verde": "cv", "Colombia": "co", "Croatia": "hr", "Curaçao": "cw",
    "Czech Republic": "cz", "DR Congo": "cd", "Ecuador": "ec", "Egypt": "eg",
    "England": "gb-eng", "France": "fr", "Germany": "de", "Ghana": "gh",
    "Haiti": "ht", "Iran": "ir", "Iraq": "iq", "Ivory Coast": "ci",
    "Japan": "jp", "Jordan": "jo", "Korea Republic": "kr", "Mexico": "mx",
    "Morocco": "ma", "Netherlands": "nl", "New Zealand": "nz", "Norway": "no",
    "Panama": "pa", "Paraguay": "py", "Portugal": "pt", "Qatar": "qa",
    "Saudi Arabia": "sa", "Scotland": "gb-sct", "Senegal": "sn", "South Africa": "za",
    "Spain": "es", "Sweden": "se", "Switzerland": "ch", "Tunisia": "tn",
    "Turkey": "tr", "United States": "us", "Uruguay": "uy", "Uzbekistan": "uz",
}

# Knockout bracket adjacency, derived from the sheet's bracket layout (cols BK..CR).
# Each entry: match number -> (round, [home_slot, away_slot]).
# Slots are either a group-position rule, a "3rd Group ..." pool rule, or a feeder
# reference: "W74" = winner of match 74, "L101" = loser of match 101.
KNOCKOUT = {
    # Round of 32
    73: ("R32", ["2A", "2B"]),
    74: ("R32", ["1E", "3rd Group A/B/C/D/F"]),
    75: ("R32", ["1F", "2C"]),
    76: ("R32", ["1C", "2F"]),
    77: ("R32", ["1I", "3rd Group C/D/F/G/H"]),
    78: ("R32", ["2E", "2I"]),
    79: ("R32", ["1A", "3rd Group C/E/F/H/I"]),
    80: ("R32", ["1L", "3rd Group E/H/I/J/K"]),
    81: ("R32", ["1D", "3rd Group B/E/F/I/J"]),
    82: ("R32", ["1G", "3rd Group A/E/H/I/J"]),
    83: ("R32", ["2K", "2L"]),
    84: ("R32", ["1H", "2J"]),
    85: ("R32", ["1B", "3rd Group E/F/G/I/J"]),
    86: ("R32", ["1J", "2H"]),
    87: ("R32", ["1K", "3rd Group D/E/I/J/L"]),
    88: ("R32", ["2D", "2G"]),
    # Round of 16
    89: ("R16", ["W74", "W77"]),
    90: ("R16", ["W73", "W75"]),
    91: ("R16", ["W76", "W78"]),
    92: ("R16", ["W79", "W80"]),
    93: ("R16", ["W83", "W84"]),
    94: ("R16", ["W81", "W82"]),
    95: ("R16", ["W86", "W88"]),
    96: ("R16", ["W85", "W87"]),
    # Quarterfinals
    97: ("QF", ["W89", "W90"]),
    98: ("QF", ["W93", "W94"]),
    99: ("QF", ["W91", "W92"]),
    100: ("QF", ["W95", "W96"]),
    # Semifinals
    101: ("SF", ["W97", "W98"]),
    102: ("SF", ["W99", "W100"]),
    # Third-place play-off
    103: ("3P", ["L101", "L102"]),
    # Final
    104: ("F", ["W101", "W102"]),
}

ROUND_NAMES = {
    "R32": "Round of 32", "R16": "Round of 16", "QF": "Quarterfinals",
    "SF": "Semi-Finals", "3P": "Third-Place Play-Off", "F": "Final",
}

# --- Venues -------------------------------------------------------------------
# football-data.org doesn't expose venues for the World Cup, so the host city per
# match is taken from the official 2026 schedule (venues are assigned by match slot
# and fixed since the schedule's release). City -> (stadium, display city name).
CITY_VENUE = {
    "Mexico City": ("Estadio Azteca", "Mexico City"),
    "Zapopan": ("Estadio Akron", "Guadalajara"),
    "Monterrey": ("Estadio BBVA", "Monterrey"),
    "Toronto": ("BMO Field", "Toronto"),
    "Vancouver": ("BC Place", "Vancouver"),
    "East Rutherford": ("MetLife Stadium", "New York/New Jersey"),
    "Inglewood": ("SoFi Stadium", "Los Angeles"),
    "Arlington": ("AT&T Stadium", "Dallas"),
    "Houston": ("NRG Stadium", "Houston"),
    "Atlanta": ("Mercedes-Benz Stadium", "Atlanta"),
    "Miami Gardens": ("Hard Rock Stadium", "Miami"),
    "Philadelphia": ("Lincoln Financial Field", "Philadelphia"),
    "Seattle": ("Lumen Field", "Seattle"),
    "Santa Clara": ("Levi's Stadium", "San Francisco Bay Area"),
    "Kansas City": ("Arrowhead Stadium", "Kansas City"),
    "Foxborough": ("Gillette Stadium", "Boston"),
}

# Team-name aliases used by the schedule source -> our seed names.
VENUE_TEAM_ALIASES = {
    "South Korea": "Korea Republic",
    "Czechia": "Czech Republic",
    "Türkiye": "Turkey",
    "USA": "United States",
}

# Group-stage host city per match, keyed by "HOME|AWAY" (order as published).
GROUP_VENUES = """\
Mexico|South Africa|Mexico City
South Korea|Czechia|Zapopan
Canada|Bosnia and Herzegovina|Toronto
USA|Paraguay|Inglewood
Qatar|Switzerland|Santa Clara
Brazil|Morocco|East Rutherford
Haiti|Scotland|Foxborough
Australia|Türkiye|Vancouver
Germany|Curaçao|Houston
Netherlands|Japan|Arlington
Ivory Coast|Ecuador|Philadelphia
Sweden|Tunisia|Monterrey
Spain|Cape Verde|Atlanta
Belgium|Egypt|Seattle
Saudi Arabia|Uruguay|Miami Gardens
Iran|New Zealand|Inglewood
France|Senegal|East Rutherford
Iraq|Norway|Foxborough
Argentina|Algeria|Kansas City
Austria|Jordan|Santa Clara
Portugal|DR Congo|Houston
England|Croatia|Arlington
Ghana|Panama|Toronto
Uzbekistan|Colombia|Mexico City
Czechia|South Africa|Atlanta
Switzerland|Bosnia and Herzegovina|Inglewood
Canada|Qatar|Vancouver
Mexico|South Korea|Zapopan
USA|Australia|Seattle
Scotland|Morocco|Foxborough
Brazil|Haiti|Philadelphia
Türkiye|Paraguay|Santa Clara
Netherlands|Sweden|Houston
Germany|Ivory Coast|Toronto
Ecuador|Curaçao|Kansas City
Tunisia|Japan|Monterrey
Spain|Saudi Arabia|Atlanta
Belgium|Iran|Inglewood
Uruguay|Cape Verde|Miami Gardens
New Zealand|Egypt|Vancouver
Argentina|Austria|Arlington
France|Iraq|Philadelphia
Norway|Senegal|East Rutherford
Jordan|Algeria|Santa Clara
Portugal|Uzbekistan|Houston
England|Ghana|Foxborough
Panama|Croatia|Toronto
Colombia|DR Congo|Zapopan
Switzerland|Canada|Vancouver
Bosnia and Herzegovina|Qatar|Seattle
Scotland|Brazil|Miami Gardens
Morocco|Haiti|Atlanta
Czechia|Mexico|Mexico City
South Africa|South Korea|Monterrey
Curaçao|Ivory Coast|Philadelphia
Ecuador|Germany|East Rutherford
Japan|Sweden|Arlington
Tunisia|Netherlands|Kansas City
Türkiye|USA|Inglewood
Paraguay|Australia|Santa Clara
Norway|France|Foxborough
Senegal|Iraq|Toronto
Cape Verde|Saudi Arabia|Houston
Uruguay|Spain|Monterrey
Egypt|Iran|Seattle
New Zealand|Belgium|Vancouver
Panama|England|East Rutherford
Croatia|Ghana|Philadelphia
Colombia|Portugal|Miami Gardens
DR Congo|Uzbekistan|Atlanta
Algeria|Austria|Kansas City
Jordan|Argentina|Arlington"""

# Knockout host city per round, in chronological (kickoff) order.
KNOCKOUT_VENUES = {
    "R32": ["Inglewood", "Houston", "Foxborough", "Monterrey", "Arlington",
            "East Rutherford", "Mexico City", "Atlanta", "Seattle", "Santa Clara",
            "Inglewood", "Toronto", "Vancouver", "Arlington", "Miami Gardens",
            "Kansas City"],
    "R16": ["Houston", "Philadelphia", "East Rutherford", "Mexico City",
            "Arlington", "Seattle", "Atlanta", "Vancouver"],
    "QF": ["Foxborough", "Inglewood", "Miami Gardens", "Kansas City"],
    "SF": ["Arlington", "Atlanta"],
    "3P": ["Miami Gardens"],
    "F": ["East Rutherford"],
}


def venue_for_city(city):
    stadium, disp = CITY_VENUE[city]
    return {"stadium": stadium, "city": disp}


def norm_venue_team(name):
    return VENUE_TEAM_ALIASES.get(name, name)


def to_utc_iso(serial):
    """Excel serial (in GMT+5:30) -> UTC ISO 8601 string, or None."""
    if serial is None:
        return None
    local_dt = from_excel(serial)
    return (local_dt - SHEET_TZ_OFFSET).strftime("%Y-%m-%dT%H:%M:%SZ")


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["2026 World Cup"]
    settings = wb["Settings"]

    # --- FIFA rank points (Settings B19:C66) ---
    rank_points = {}
    for r in range(19, 67):
        name = settings.cell(row=r, column=2).value
        pts = settings.cell(row=r, column=3).value
        if name and isinstance(pts, (int, float)):
            rank_points[str(name).strip()] = pts

    # --- Group-stage fixtures (rows 7..78: match # in col A, teams in E/H, UTC in R) ---
    fixtures = []
    for r in range(7, 79):
        num = ws.cell(row=r, column=1).value
        home = ws.cell(row=r, column=5).value
        away = ws.cell(row=r, column=8).value
        serial = ws.cell(row=r, column=18).value  # col R = kickoff datetime
        if not (isinstance(num, (int, float)) and home and away):
            continue
        fixtures.append({
            "id": int(num),
            "stage": "group",
            "home": str(home).strip(),
            "away": str(away).strip(),
            "kickoff": to_utc_iso(serial),
        })

    # --- Groups + their teams (standings tables at cols AD/AE, headers every 6 rows) ---
    # The group label is in col J on the header row; teams follow in col AE for 4 rows.
    groups = {}
    for r in range(8, 79):
        label = ws.cell(row=r, column=10).value  # col J
        if isinstance(label, str) and label.startswith("Group "):
            g = label.replace("Group ", "").strip()
            teams = []
            for k in range(r, r + 4):
                t = ws.cell(row=k, column=31).value  # col AE
                if t:
                    teams.append(str(t).strip())
            groups[g] = teams

    # --- Knockout bracket: pull kickoff per match number from the sheet ---
    ko_times = {}
    # match numbers live in col BK (63) for R32, BR (70) R16, BY (77) QF, CF (84) SF, CM (91) Final/3P
    # the kickoff datetime sits in a nearby column on the row above the match number.
    for r in range(6, 112):
        for c in (63, 70, 77, 84, 91):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, (int, float)) and 73 <= v <= 104:
                # search the small block around this cell for a datetime serial
                for rr in range(r - 2, r + 1):
                    for cc in range(c, c + 10):
                        dv = ws.cell(row=rr, column=cc).value
                        if isinstance(dv, (int, float)) and 46000 < dv < 47000:
                            ko_times[int(v)] = to_utc_iso(dv)
                            break
                    if int(v) in ko_times:
                        break

    knockout = []
    for num in sorted(KNOCKOUT):
        rnd, slots = KNOCKOUT[num]
        knockout.append({
            "id": num,
            "stage": "knockout",
            "round": rnd,
            "roundName": ROUND_NAMES[rnd],
            "homeSlot": slots[0],
            "awaySlot": slots[1],
            "kickoff": ko_times.get(num),
        })

    # --- Assign venues -------------------------------------------------------
    # Group stage: match the published "home|away|city" rows to our fixtures by
    # unordered team pair (some published orders are reversed vs the seed).
    group_city = {}
    for line in GROUP_VENUES.splitlines():
        home, away, city = line.split("|")
        key = frozenset({norm_venue_team(home), norm_venue_team(away)})
        group_city[key] = city
    missing_venue = []
    for fx in fixtures:
        city = group_city.get(frozenset({fx["home"], fx["away"]}))
        if city:
            fx["venue"] = venue_for_city(city)
        else:
            missing_venue.append(fx["id"])

    # Knockout: zip each round's published cities (chronological) onto our knockout
    # matches sorted by kickoff within that round.
    for rnd, cities in KNOCKOUT_VENUES.items():
        games = sorted(
            (k for k in knockout if k["round"] == rnd),
            key=lambda k: k["kickoff"] or "",
        )
        for game, city in zip(games, cities):
            game["venue"] = venue_for_city(city)

    teams = sorted(groups_to_teams(groups))
    seed = {
        "tournament": "FIFA World Cup 2026",
        "source": "fifa_2026.xlsx",
        "teams": [
            {"name": t, "flag": FLAG_CODES.get(t), "rankPoints": rank_points.get(t)}
            for t in teams
        ],
        "groups": groups,
        "fixtures": fixtures,
        "knockout": knockout,
    }

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(seed, f, ensure_ascii=False, indent=2)

    missing = [t for t in teams if t not in FLAG_CODES]
    venued = sum(1 for f in fixtures + knockout if f.get("venue"))
    print(f"Wrote {OUT}")
    print(f"  teams={len(teams)} groups={len(groups)} "
          f"group_fixtures={len(fixtures)} knockout={len(knockout)}")
    print(f"  venues assigned: {venued}/{len(fixtures) + len(knockout)}")
    if missing:
        print(f"  WARNING missing flag codes: {missing}")
    if missing_venue:
        print(f"  WARNING group fixtures without venue: {missing_venue}")


def groups_to_teams(groups):
    out = set()
    for ts in groups.values():
        out.update(ts)
    return out


if __name__ == "__main__":
    main()
